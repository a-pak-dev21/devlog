"""
Script to automate the cleaning/modifying guest's data in Excel format.
See guest_cleaner_README.md for detailed specification and usage.
"""

import pandas as pd
from pandas import isna
from pathlib import Path
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from random import choice
from pet_projects.excel_report_refiner.logs.loggers import countries_logger


class DataCleaner:

    # Initializing path to our data_file and new instance of DataFrame to work with data
    # Also control if both values are strings

    def __init__(self, file_name: str, sheet_name: str, viza_path: str | Path, addresses_path: str | Path) -> None:
        assert isinstance(file_name, str), "The name of file should be a string value"
        assert isinstance(sheet_name, str), "Insert valid name of the excel sheet"
        self.file_dir = Path("draft") / file_name
        self.df = pd.read_excel(self.file_dir, sheet_name=sheet_name)
        self.viza_obligated = self._load_viza_required(viza_path)
        self.addresses = self._load_addresses(addresses_path)
        self.wb = load_workbook(self.file_dir, keep_vba=True)
        self.ws = self.wb[sheet_name]

    # From this part, we will start to modify our table
    # First in priorities we will perform all method which removing
    # invalid rows to not fill and process them for nothing
    # after modify and fill remaining rows

    @staticmethod
    def _load_viza_required(path: str | Path) -> set:
        return set(pd.read_csv(path)["Visa Obligated"])

    @staticmethod
    def _load_addresses(path: str | Path) -> json:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def remove_invalid_dob(self) -> None:
        if not isinstance(self.df["datum narození"].iloc[0], str):
            self.df["datum narození"] = self.df["datum narození"].dt.strftime("%d.%m.%Y")

        self.df = self.df[self.df["datum narození"] != "00.00.0000"].copy()

    def remove_duplicates(self) -> None:
        self.df = self.df.drop_duplicates(subset=["příjmení", "jméno", "číslo cestovního dokladu"])

    @staticmethod
    def _check_first_name(first_name: str) -> str:
        split_name: list[str] = first_name.split(" ")
        return " ".join(split_name[:2]) if len(split_name) > 2 else first_name

    def check_first_name(self) -> None:
        self.df["jméno"] = self.df["jméno"].apply(self._check_first_name)
        blank_name = self.df[(self.df["jméno"].isna()) & (self.df["příjmení"].str.split(" ").str.len() >= 2)]
        for idx, row in blank_name[["příjmení", "jméno"]].iterrows():
            split_name = row["příjmení"].split(" ")
            last_name, first_name = " ".join(split_name[:-1]), split_name[-1]
            self.df.loc[idx, ["příjmení", "jméno"]] = last_name, first_name
        self.df = self.df.dropna(subset=["jméno"])

    def viza_validation(self) -> None:
        self.df = self.df[(~self.df["státní občanství"].isin(self.viza_obligated))
                          | ((self.df["státní občanství"].isin(self.viza_obligated)) & (self.df["číslo víza"].notna()))]
        self.df["číslo víza"] = self.df["číslo víza"].fillna("")

    def pass_filter(self) -> None:
        self.df = self.df.dropna(subset=["číslo cestovního dokladu"])
        self.df = self.df[self.df["číslo cestovního dokladu"].str.len() > 5]
        self.df = self.df[~self.df["číslo cestovního dokladu"].str.startswith("XXXX")]

    def fill_rsn_of_stay(self) -> None:
        self.df["Unnamed: 13"] = self.df["Unnamed: 13"].fillna("10")

    def _address_filter(self, row: pd.Series) -> str | None:
        address = row[","]
        nationality = row["státní občanství"]
        if not isinstance(address, str) or len(address) < 5:
            if row["státní občanství"] in self.addresses:
                return choice(self.addresses[nationality])
            else:
                countries_logger.info(f"The Country with ISO3: {nationality} is not in list of countries")
                return f"{address} DOPLNIT!!!"
        return address

    def address_filter(self) -> None:
        self.df[","] = self.df.apply(self._address_filter, axis=1)
        # lambda row: choice(self.addresses[row["státní občanství"]]) if (len(str(row[","])) < 5) and
        #             (row["státní občanství"] in self.addresses) else row[","]

    @staticmethod
    def _address_cleaner(address: str) -> str:
        parts: list[str] = [part.strip() for part in address.split(", ")]
        if len(parts) != 3:
            return address

        iso2, first_part, second_part = parts
        if first_part.capitalize() in ["str", ".", "Street 1"] or first_part.capitalize() == second_part.capitalize():
            return f"{iso2}, {second_part}"

        return address

    def address_cleaner(self) -> None:
        self.df[","] = self.df[","].apply(self._address_cleaner)

    def apply_changes(self) -> pd.DataFrame:
        self.remove_invalid_dob()
        self.remove_duplicates()
        self.pass_filter()
        self.check_first_name()
        self.viza_validation()
        self.fill_rsn_of_stay()
        self.address_filter()
        self.address_cleaner()
        return self.df

    def clean_table(self) -> None:
        for row in self.ws.iter_rows(
                min_row=2, max_row=self.ws.max_row, min_col=2, max_col=self.ws.max_column):
            for cell in row:
                cell.value = None

    def write_dataframe(self) -> None:
        for row_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=False), start=2):
            for col_idx, col in enumerate(row, start=1):
                value = None if isna(col) else col
                self.ws.cell(row=row_idx, column=col_idx, value=value)

    def remove_blank_rows(self) -> None:
        last_row = self.ws.max_row
        while last_row > 1:
            if self.ws[f"B{last_row}"].value in (None, ""):
                self.ws.delete_rows(last_row)
            else:
                break
            last_row -= 1

    def completion(self) -> None:
        self.apply_changes()
        self.clean_table()
        self.write_dataframe()
        self.remove_blank_rows()
        self.wb.save(self.file_dir)


csv_path = Path("data/visa_obligated.csv")
json_path = Path("data/addresses_data.json")

my_DataCleaner = DataCleaner("Ubydata_19B_copy.xlsm", "Seznam", csv_path, json_path)
my_DataCleaner.completion()
