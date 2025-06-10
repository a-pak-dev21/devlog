from openpyxl import Workbook, load_workbook
from pathlib import Path
from collections import defaultdict


def solution(data: list[list[str | int]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Sales"
    fieldnames:list[str] = ["Date", "Product", "Amount"]
    ws.append(fieldnames)
    for sale in data:
        ws.append(sale)

    file_dir = Path("datas") / "sales.xlsx"
    wb.save(file_dir)

    wb_r = load_workbook(file_dir)
    ws_r = wb_r.active
    products = defaultdict(int)
    sales_by_day = defaultdict(int)
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        products[row[1]] += row[2]
        sales_by_day[row[0]] += row[2]

    best_sales_day = max(sales_by_day.items(), key=lambda x: x[1])
    products_revenue = [[key, val] for key, val in products.items()]

    summary_dir = Path("datas") / "sales_summary.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Sales Summary"
    ws2.append(["Product Name", "Products' revenue"])
    total_revenue = 0
    for product in products_revenue:
        ws2.append(product)
        total_revenue += product[1]
    ws2.append(["Total Revenue: ", total_revenue])
    ws2.append(["Most profitable day:"])
    ws2.append([best_sales_day[0][0], best_sales_day[0][1]])
    wb2.save(summary_dir)


rows = [
    ["2024-01-01", "Book", 15],
    ["2024-01-01", "Pen", 5],
    ["2024-01-02", "Notebook", 20],
    ["2024-01-03", "Book", 30],
    ["2024-01-03", "Pencil", 3],
    ["2024-01-04", "Eraser", 4],
    ["2024-01-04", "Book", 25],
    ["2024-01-05", "Notebook", 10],
    ["2024-01-05", "Pen", 8],
    ["2024-01-06", "Marker", 12],
    ["2024-01-06", "Book", 18]
]

solution(rows)
