from openpyxl import Workbook, load_workbook
from pathlib import Path


def solution():
    wb = Workbook()
    ws = wb.active
    ws.title = "Students and grades"
    ws.append(["Student", "Grade"])
    ws.append(["Alice", 4.4])
    ws.append(["Jeremy", 1.9])
    ws.append(["Jane", 1.5])
    ws.append(["Clark", 3.3])
    ws.append(["Bruce", 2.1])
    ws.append(["Selene", 4.0])
    ws.append(["Angelo", 3.5])

    folder_dir = Path("datas")
    folder_dir.mkdir(exist_ok=True)
    file_dir = folder_dir / "grades.xlsx"

    wb.save(file_dir)

    wb_r = load_workbook(file_dir)
    ws_r = wb_r.active

    grades = [row[1] for row in ws_r.iter_rows(min_row=2, values_only=True)]
    avg_grade = round(sum(grades) / len(grades), 2)
    print(avg_grade)
    ws_r.append(["Average", avg_grade])
    wb_r.save(file_dir)


solution()
